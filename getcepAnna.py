import streamlit as st
import datetime
import os
import openpyxl
import requests
import locale
import time


# Função para carregar a chave de API do arquivo externo
def load_api_key(file_path):
    with open(file_path, "r") as key_file:
        return key_file.read().strip()


# Função para obter o CEP usando a API do Google Geocoding
def get_cep_from_address(address, city, state, api_key):
    full_address = f"{address}, {city}, {state}, Brasil"

    base_url = "https://maps.googleapis.com/maps/api/geocode/json"
    params = {
        "address": full_address,
        "key": api_key,
    }

    response = requests.get(base_url, params=params)
    data = response.json()

    if data["status"] == "OK":
        results = data["results"]
        if results:
            for component in results[0]["address_components"]:
                if "postal_code" in component["types"]:
                    return component["long_name"]

    return None


# Função para imprimir informações sobre o registro sendo tratado
def print_processing_info(record_number, total_records):
    print(f"Processando registro {record_number} de {total_records}")


# Função para obter informações do arquivo
def get_file_info(file_name):
    try:
        file_path = os.path.join(os.getcwd(), file_name)
        file_size = os.path.getsize(file_path)
        creation_time = datetime.datetime.fromtimestamp(os.path.getctime(file_path))
        return {
            "file_size": file_size,
            "creation_time": creation_time,
            "file_path": file_path
        }
    except Exception as e:
        return {
            "error": str(e)
        }


# Função para contar as linhas do arquivo Excel
def count_excel_rows(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        row_count = sheet.max_row - 1  # Exclude header row
        workbook.close()
        return row_count
    except Exception as e:
        return {
            "error": str(e)
        }

st.set_page_config(
    page_title="Get CEP",
    page_icon="🏙️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.locx-getcep.com/help',
        'Report a bug': "https://www.locx-getcep.com/bug",
        'About': "# Get CEP"
        }
)

enable_bt_upload = False
    
with open('style.css') as f:
    st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)

    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')


st.markdown("""
# 📝 :orange[GetCEP] API Google
:orange[_Versão_ _1.5.010_ _15/08/20023_]

Retorna o CEP a partir de um endereço passado como parãmetro.                      
1. Usa uma planilha Excel que contêm os endereços
2. Necessita consumir a API do Google, portanto é necessário a chave de utilização
3. Retorna os CEPs encontrados numa planilha Excel
"""
)

st.divider()

col1, col2 = st.columns(2)
with st.container():        # upload_resume
    with col1:
        res_file = st.file_uploader('📁 Upload sua planilha de endereços em Excel',
                                ['xlsx','xls'],
                                accept_multiple_files=False,
                                help='Selecione o arquivo Excel com os endereços.',
                                disabled=enable_bt_upload)
    with col2:
        st.markdown("""Arquivo Excel com os dados dos endereços para serem usados com a API do Google.
                    
                    Este arquivo deve ter as seguintes colunas:
                    
        1.COD.AUXILIAR / 2.UF / 3.CIDADE / 4.LOGRADOURO / 5.NUMERO
                    """)

    if res_file:
        st.toast('Arquivo selecionado com sucesso!')

        with st.container():
            file_info = get_file_info(res_file.name)    # Obter informações do arquivo
            col1, col2, col3 = st.columns([1, 1, 3])
            if "error" not in file_info:
                formatted_creation_time = file_info['creation_time'].strftime('%d/%m/%Y %H:%M:%S')
                with col1:
                    st.write(f":orange[Tamanho do arquivo:] {file_info['file_size']} bytes")
                with col2:
                    st.write(f":orange[Data de criação:] {formatted_creation_time}")
                with col3:
                    st.write(f":orange[Caminho completo:] {file_info['file_path']}")
            
                excel_row_count = count_excel_rows(res_file.name)   # Contar as linhas do arquivo Excel
                if isinstance(excel_row_count, int):
                    st.markdown(f"##### :orange[Número de CEPs neste arquivo Excel: {excel_row_count}]")
                else:
                    st.write(f"Erro ao contar linhas do arquivo Excel: {excel_row_count.get('error', 'Erro desconhecido')}")
            else:
                st.write(f"Erro ao obter informações do arquivo: {file_info['error']}")
            
            st.divider()
        
with st.form('meu_form'):
    output_file = st.text_input('Nome do arquivo a ser salvo',
                                        value='resultado_ceps.xlsx')

    execute_button = st.form_submit_button(label='Executar')   # Botão de "Executar"
    if execute_button and res_file:
        with st.spinner(text='Aguarde...'):
            time.sleep(2)
            
            api_key_file = "C:/python/projetos_locx/API/google/api_key.txt"
            api_key = load_api_key(api_key_file)

            # Definir o nome da planilha Excel e as colunas de endereço, cidade e unidade federativa (U.F.)
            excel_file = res_file 
            id_column = 0           # Coluna COD. AUXILIAR
            state_column = 1        # Coluna U.F.
            city_column = 2         # Coluna Cidade
            address_column = 3      # Coluna Endereço
            numero_column = 4       # Coluna Numero

            workbook = openpyxl.load_workbook(excel_file) # Carregar a planilha
            sheet = workbook.active

            # Criar novo arquivo Excel para gravar os resultados
            output_workbook = openpyxl.Workbook()
            output_sheet = output_workbook.active

            # Definir cabeçalhos para a nova planilha de saída
            output_sheet.append(["Cod.Aux", "Endereço", "Cidade", "U.F.", "CEP"])

            # Percorrer as linhas da planilha e obter os CEPs
            total_records = sheet.max_row - 1
            processing_info = []
            
            for record_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True, max_col=7), start=1):
                id_aux = row[id_column]
                address = f"{row[address_column]}, {row[numero_column]}"
                city = row[city_column]
                state = row[state_column]

                if address and city and state:
                    processing_info.append([id_aux, address, city, state, "Processando..."])
                    #print_processing_info(record_number, total_records)
                    cep = get_cep_from_address(address, city, state, api_key)
                    if cep:
                        output_sheet.append([id_aux, address, city, state, cep])
                        processing_info[-1][-1] = cep
                    else:
                        output_sheet.append([id_aux, address, city, state, "Não encontrado"])
                        processing_info[-1][-1] = "Não encontrado"
                else:
                    output_sheet.append(["Dados incompletos", "", "", "", ""])
                    processing_info.append(["Dados incompletos", "", "", "", ""])

            output_workbook.save(output_file) # Salvar a nova planilha de saída
            
            workbook.close()
            output_workbook.close()

            st.warning(f"Resultados gravados em '{output_file}'")
            st.table(processing_info)

if execute_button:
    if os.path.exists(output_file):
        download_button_label = f'Clique para baixar {output_file}'
        with open(output_file, "rb") as f:
            output_file_data = f.read()
            
        st.download_button(
            label=download_button_label,
            data=output_file_data,
            file_name=output_file,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        
